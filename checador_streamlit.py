"""
Checador de Facturas — D.M. Mexicana SA de CV
Streamlit app: lee un Excel de ventas, busca cada # de venta en las facturas PDF
de Google Drive y extrae Folio, Descripción y Cantidad.
"""

import io
import json
import re
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path

import openpyxl
import pandas as pd
import pdfplumber
import streamlit as st
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
from openpyxl.styles import Alignment, Font, PatternFill

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

DEFAULT_FOLDER_URL = (
    "https://drive.google.com/drive/folders/"
    "1_rl-CaNJDAQHl0-RAMm99aybPJlM3368?usp=drive_link"
)
SCOPES = ["https://www.googleapis.com/auth/drive.readonly"]
SCRIPT_DIR = Path(__file__).parent


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def extract_folder_id(url_or_id: str) -> str:
    m = re.search(r"/folders/([a-zA-Z0-9_-]+)", url_or_id)
    return m.group(1) if m else url_or_id.strip()


def find_local_service_account() -> dict | None:
    """Return the first service-account JSON found next to this script."""
    for f in SCRIPT_DIR.glob("*.json"):
        try:
            data = json.loads(f.read_text(encoding="utf-8"))
            if data.get("type") == "service_account":
                return data
        except Exception:
            pass
    return None


def load_credentials() -> dict | None:
    """
    Priority:
      1. Streamlit Cloud secrets  (key: gcp_service_account)
      2. Local service-account JSON in the same folder as this script
    Returns the parsed JSON dict or None.
    """
    try:
        sa = st.secrets.get("gcp_service_account")
        if sa:
            return dict(sa)
    except Exception:
        pass
    return find_local_service_account()


@st.cache_resource(show_spinner=False)
def build_drive_service(json_str: str):
    """Build and cache a Google Drive API service from a JSON string."""
    info = json.loads(json_str)
    creds = service_account.Credentials.from_service_account_info(info, scopes=SCOPES)
    return build("drive", "v3", credentials=creds)


# ---------------------------------------------------------------------------
# Invoice index
# ---------------------------------------------------------------------------

class InvoiceIndex:
    """Maps MercadoLibre sale numbers → {folio, descripcion, cant, file}."""

    def __init__(self):
        self._data: dict[str, dict] = {}
        self._lock = threading.Lock()

    # ---- public API --------------------------------------------------------

    def add_pdf(self, text: str, file_name: str) -> list[str]:
        sale_numbers = self._find_sale_numbers(text)
        if not sale_numbers:
            return []
        entry = {
            "folio":       self._extract_folio(text),
            "descripcion": self._extract_descripcion(text),
            "cant":        self._extract_cant(text),
            "file":        file_name,
        }
        with self._lock:
            for sn in sale_numbers:
                self._data[sn] = entry
        return sale_numbers

    def lookup(self, sale_number: str) -> dict | None:
        return self._data.get(str(sale_number).strip())

    def __len__(self) -> int:
        return len(self._data)

    # ---- extractors --------------------------------------------------------

    @staticmethod
    def _find_sale_numbers(text: str) -> list[str]:
        patterns = [
            r"Venta\s+DM\s+Mercado\s*[Ll]ibre\s*[-–]\s*(\d{10,})",
            r"Mercado\s*[Ll]ibre\s*[-–]\s*(\d{10,})",
            r"\b(2\d{15})\b",   # ML orders: 16-digit numbers starting with 2
        ]
        found: set[str] = set()
        for pat in patterns:
            found.update(re.findall(pat, text, re.IGNORECASE))
        return list(found)

    @staticmethod
    def _extract_folio(text: str) -> str:
        for m in re.finditer(r"\bFOLIO\b\s*[:\s]+([A-Z0-9-]{1,40})", text, re.IGNORECASE):
            val = m.group(1).strip()
            if re.match(r"[0-9A-F]{8}-", val, re.IGNORECASE):
                continue          # skip FOLIO FISCAL (UUID)
            if re.match(r"\d+$", val):
                return val
        return ""

    @staticmethod
    def _extract_descripcion(text: str) -> str:
        # Pattern: D######  CUM  CPS(8+)  DESCRIPTION  quantity
        m = re.search(
            r"D\d{5,}\s+\S+\s+\d{6,}\s+([\w\s,/().ÁÉÍÓÚÜÑ]+?)"
            r"(?=\s+\d+\.\d{2}\s+\$|\n)",
            text, re.IGNORECASE,
        )
        if m:
            return m.group(1).strip()

        product_kw = re.compile(
            r"FORMULA|LECHE|LACTANTE|CRECE|BEBE|CABRA|VACA|NUTRI|INFAN|NIÑ|"
            r"VITAMINA|CEREAL|ALIMENTO|SUPLEMENTO|COMPLEMENTO|[0-9]+G\b",
            re.IGNORECASE,
        )
        skip_kw = re.compile(
            r"\bFOLIO\b|\bFECHA\b|\bRECEPTOR\b|\bRFC\b|\bLUGAR\b|\bEMISOR\b"
            r"|\bMONEDA\b|\bSUCURSAL\b|\bCODIGO\b|\bPRECIO\b|\bIMPORTE\b"
            r"|\bTOTAL\b|\bSUBTOTAL\b|\bIVA\b|\bTIPO\b|\bSERIE\b"
            r"|\bREFERENCIA\b|\bPAGO\b|\bFISCAL\b|\bEFECTO\b|\bINGRESO\b",
            re.IGNORECASE,
        )
        for line in text.splitlines():
            line = line.strip()
            if len(line) < 15 or skip_kw.search(line):
                continue
            if product_kw.search(line):
                return line
        return ""

    @staticmethod
    def _extract_cant(text: str) -> str:
        m = re.search(r"\b(\d{1,4}\.\d{2})\s+\$[\d,]+\.\d{2,4}", text)
        if m and float(m.group(1)) < 100_000:
            return m.group(1)
        return ""


# ---------------------------------------------------------------------------
# Drive helpers
# ---------------------------------------------------------------------------

def list_all_pdfs(service, folder_id: str) -> list[dict]:
    files, token = [], None
    while True:
        r = service.files().list(
            q=f"'{folder_id}' in parents and mimeType='application/pdf' and trashed=false",
            fields="nextPageToken, files(id, name)",
            pageSize=1000,
            pageToken=token,
        ).execute()
        files.extend(r.get("files", []))
        token = r.get("nextPageToken")
        if not token:
            break
    return files


def download_pdf(service, file_id: str) -> io.BytesIO:
    req = service.files().get_media(fileId=file_id)
    buf = io.BytesIO()
    dl = MediaIoBaseDownload(buf, req)
    done = False
    while not done:
        _, done = dl.next_chunk()
    buf.seek(0)
    return buf


MAX_WORKERS = 6   # parallel PDF downloads; raise if you have many invoices


def build_index(service, folder_id: str) -> InvoiceIndex | None:
    """Download PDFs in parallel and build the invoice index with live progress."""
    files = list_all_pdfs(service, folder_id)
    if not files:
        st.warning("No se encontraron archivos PDF en la carpeta de Google Drive.")
        return None

    total = len(files)
    index = InvoiceIndex()

    # ── UI placeholders ─────────────────────────────────────────────────────
    st.caption(f"Descargando **{total}** facturas con **{MAX_WORKERS}** hilos en paralelo…")
    bar        = st.progress(0.0)
    cols       = st.columns(3)
    cnt_pdfs   = cols[0].empty()   # PDFs leídos
    cnt_ventas = cols[1].empty()   # ventas indexadas
    cnt_errors = cols[2].empty()   # errores
    log        = st.empty()

    completed = 0
    errors    = 0
    msgs: list[str] = []

    def _process(f: dict) -> tuple[str, list[str]]:
        """Download + extract text + index one PDF. Runs in worker thread."""
        buf  = download_pdf(service, f["id"])
        text = "\n".join(p.extract_text() or "" for p in pdfplumber.open(buf).pages)
        sns  = index.add_pdf(text, f["name"])
        return f["name"], sns

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as pool:
        futures = {pool.submit(_process, f): f for f in files}

        for future in as_completed(futures):
            completed += 1
            try:
                name, sns = future.result()
                icon = "✅" if sns else "⬜"
                msgs.append(f"{icon} {name}" + (f"  →  {', '.join(sns)}" if sns else ""))
            except Exception as exc:
                errors += 1
                msgs.append(f"⚠️  {futures[future]['name']}: {exc}")

            # Update UI (runs in the main Streamlit thread via as_completed iteration)
            pct = completed / total
            bar.progress(pct, text=f"{completed} / {total} PDFs  ({pct:.0%})")
            cnt_pdfs.metric("📄 PDFs leídos",       f"{completed} / {total}")
            cnt_ventas.metric("🗂️ Ventas indexadas", len(index))
            cnt_errors.metric("⚠️ Errores",          errors)
            log.code("\n".join(msgs[-40:]))   # rolling log, last 40 lines

    bar.progress(1.0, text=f"Indexación completa ✓  —  {len(index)} ventas en {total} facturas")
    return index


# ---------------------------------------------------------------------------
# Excel processing
# ---------------------------------------------------------------------------

def read_excel_sales(raw_bytes: bytes) -> tuple[int, int, list[str]] | tuple[None, None, None]:
    """Return (header_row_0idx, sale_col_0idx, sale_numbers_list)."""
    df = pd.read_excel(io.BytesIO(raw_bytes), header=None, dtype=str)
    hr, sc = None, None
    for ri, row in df.iterrows():
        for ci, cell in enumerate(row):
            if cell and "# de venta" in str(cell).lower():
                hr, sc = int(ri), int(ci)
                break
        if hr is not None:
            break

    if hr is None:
        st.error("No se encontró la columna **'# de venta'** en el Excel.")
        return None, None, None

    numbers = []
    for v in df.iloc[hr + 1:, sc]:
        v = str(v).strip()
        if not v or v.lower() == "nan":
            continue
        if re.match(r"^\d+\.0$", v):
            v = v[:-2]
        numbers.append(v)

    return hr, sc, numbers


def match_sales(sale_numbers: list[str], index: InvoiceIndex) -> list[dict]:
    bar  = st.progress(0.0, text="Buscando coincidencias…")
    rows = []
    for i, sn in enumerate(sale_numbers, start=1):
        bar.progress(i / len(sale_numbers), text=f"[{i}/{len(sale_numbers)}] {sn}")
        entry = index.lookup(sn)
        if entry:
            rows.append({
                "sale_number": sn,
                "folio":       entry["folio"],
                "descripcion": entry["descripcion"],
                "cant":        entry["cant"],
                "found":       True,
            })
        else:
            rows.append({
                "sale_number": sn,
                "folio":       "NO ENCONTRADO",
                "descripcion": "",
                "cant":        "",
                "found":       False,
            })
    bar.progress(1.0, text="Búsqueda completa ✓")
    return rows


def create_output_excel(
    original_bytes: bytes,
    header_row_0: int,
    sale_col_0: int,
    results: list[dict],
) -> bytes:
    wb = openpyxl.load_workbook(io.BytesIO(original_bytes))
    ws = wb.active

    hdr_row  = header_row_0 + 1   # openpyxl is 1-indexed
    sale_col = sale_col_0 + 1

    ws.insert_cols(sale_col + 1, 3)

    gold   = PatternFill("solid", fgColor="FFD700")
    bold   = Font(bold=True, name="Calibri", size=11)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for offset, title in enumerate(["Folio", "Descripción", "Cant."], start=1):
        c = ws.cell(row=hdr_row, column=sale_col + offset, value=title)
        c.font, c.fill, c.alignment = bold, gold, center

    green = PatternFill("solid", fgColor="E8F5E9")
    red   = PatternFill("solid", fgColor="FFEBEE")
    lookup = {r["sale_number"]: r for r in results}

    for row_n in range(hdr_row + 1, ws.max_row + 1):
        raw = ws.cell(row=row_n, column=sale_col).value
        if raw is None:
            continue
        key = str(raw).strip()
        if key.endswith(".0"):
            key = key[:-2]
        if key not in lookup:
            continue
        r    = lookup[key]
        fill = green if r["found"] else red
        for offset, field in enumerate(["folio", "descripcion", "cant"], start=1):
            c = ws.cell(row=row_n, column=sale_col + offset, value=r[field])
            c.fill = fill

    col_letter = lambda col: ws.cell(row=1, column=col).column_letter
    ws.column_dimensions[col_letter(sale_col + 1)].width = 12
    ws.column_dimensions[col_letter(sale_col + 2)].width = 65
    ws.column_dimensions[col_letter(sale_col + 3)].width = 8

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Streamlit UI
# ---------------------------------------------------------------------------

def main() -> None:
    st.set_page_config(
        page_title="Checador de Facturas — D.M. Mexicana",
        page_icon="🧾",
        layout="wide",
    )

    # ── Session state defaults ──────────────────────────────────────────────
    for key in ("service", "folder_id", "invoice_index"):
        if key not in st.session_state:
            st.session_state[key] = None

    # ── Sidebar ─────────────────────────────────────────────────────────────
    with st.sidebar:
        st.title("🧾 D.M. Mexicana")
        st.caption("Checador de Facturas")
        st.divider()

        # Credentials
        st.subheader("🔑 Credenciales")
        creds_dict = load_credentials()
        if creds_dict:
            sa_email = creds_dict.get("client_email", "desconocido")
            st.success(f"Service Account detectado  \n`{sa_email}`")
        else:
            st.info("Pega el contenido del JSON de Service Account")
            raw_json = st.text_area("JSON de Service Account", height=120, label_visibility="collapsed")
            if raw_json.strip():
                try:
                    creds_dict = json.loads(raw_json)
                    if creds_dict.get("type") != "service_account":
                        st.error("El JSON no es de tipo service_account.")
                        creds_dict = None
                except Exception:
                    st.error("JSON inválido.")
                    creds_dict = None

        st.divider()

        # Folder URL
        st.subheader("📁 Google Drive")
        folder_url = st.text_input("URL de la carpeta", value=DEFAULT_FOLDER_URL)
        folder_id  = extract_folder_id(folder_url)
        st.caption(f"ID: `{folder_id}`")

        # Connect button
        connect_disabled = creds_dict is None
        if st.button("🔌 Conectar / Reconectar", disabled=connect_disabled, use_container_width=True):
            with st.spinner("Conectando a Google Drive…"):
                try:
                    svc = build_drive_service(json.dumps(creds_dict))
                    # Quick test
                    svc.files().list(
                        q=f"'{folder_id}' in parents and trashed=false",
                        fields="files(id)", pageSize=1
                    ).execute()
                    st.session_state.service   = svc
                    st.session_state.folder_id = folder_id
                    st.session_state.invoice_index = None
                    st.success("Conectado ✅")
                except Exception as exc:
                    st.error(f"Error: {exc}")

        if st.session_state.service:
            st.success("🟢 Conectado")
            if st.button("🔄 Reindexar facturas", use_container_width=True):
                st.session_state.invoice_index = None
                st.toast("Índice limpiado — se reconstruirá al procesar.", icon="🗑️")
        else:
            st.warning("🔴 No conectado")

        st.divider()
        st.caption(
            "**Nota:** el Service Account debe tener acceso a la carpeta de Drive.  \n"
            f"Comparte la carpeta con:  \n"
            f"`{creds_dict.get('client_email', '…') if creds_dict else '…'}`"
        )

    # ── Main area ───────────────────────────────────────────────────────────
    st.title("Checador de Facturas")
    st.caption("Lee el # de venta del Excel, busca en las facturas PDF de Google Drive y extrae Folio, Descripción y Cant.")

    if not st.session_state.service:
        st.info("👈 Configura y conecta a Google Drive en la barra lateral para comenzar.")
        st.stop()

    # Excel uploader
    excel_file = st.file_uploader(
        "📊 Sube tu archivo Excel de ventas",
        type=["xlsx", "xls"],
        help="Debe contener una columna llamada '# de venta'",
    )
    if not excel_file:
        st.info("Sube un archivo Excel para continuar.")
        st.stop()

    # Quick preview of the Excel
    with st.expander("Vista previa del Excel", expanded=False):
        try:
            preview = pd.read_excel(io.BytesIO(excel_file.read()), header=1, nrows=5)
            st.dataframe(preview, use_container_width=True)
            excel_file.seek(0)
        except Exception:
            excel_file.seek(0)

    # Action buttons
    st.divider()
    col_a, col_b, _ = st.columns([1, 1, 2])
    run_test = col_a.button("🔍 Probar primeros 10", use_container_width=True)
    run_all  = col_b.button("▶️ Procesar Todo",       use_container_width=True, type="primary")

    if not run_test and not run_all:
        st.stop()

    # ── Processing ──────────────────────────────────────────────────────────
    test_mode    = run_test
    excel_bytes  = excel_file.read()

    # Step 1 – read Excel
    with st.status("📖 Leyendo Excel…", expanded=False) as s:
        hr, sc, sale_numbers = read_excel_sales(excel_bytes)
        if sale_numbers is None:
            st.stop()
        s.update(label=f"Excel leído — {len(sale_numbers)} ventas encontradas", state="complete")

    if test_mode:
        sale_numbers = sale_numbers[:10]
        st.info(f"Modo prueba: procesando los primeros **{len(sale_numbers)}** números de venta.")

    # Step 2 – build index
    if st.session_state.invoice_index is None:
        with st.status("📥 Descargando e indexando facturas de Google Drive…", expanded=True) as s:
            idx = build_index(st.session_state.service, st.session_state.folder_id)
            if idx is None:
                st.stop()
            st.session_state.invoice_index = idx
            s.update(label=f"Índice listo — {len(idx)} ventas indexadas", state="complete")
    else:
        st.info(f"Usando índice en caché ({len(st.session_state.invoice_index)} ventas).")

    # Step 3 – match
    with st.status("🔎 Buscando coincidencias…", expanded=True) as s:
        results = match_sales(sale_numbers, st.session_state.invoice_index)
        found   = sum(1 for r in results if r["found"])
        s.update(label=f"Búsqueda completa — {found}/{len(results)} encontradas", state="complete")

    # Step 4 – output Excel
    output_bytes = create_output_excel(excel_bytes, hr, sc, results)

    # ── Results ─────────────────────────────────────────────────────────────
    st.divider()
    st.subheader("Resultados")

    m1, m2, m3 = st.columns(3)
    m1.metric("Total procesadas",   len(results))
    m2.metric("✅ Encontradas",     found)
    m3.metric("❌ No encontradas",  len(results) - found)

    df_res = pd.DataFrame(results)
    df_res["estado"] = df_res["found"].map({True: "✅ Encontrado", False: "❌ No encontrado"})
    st.dataframe(
        df_res[["sale_number", "folio", "descripcion", "cant", "estado"]].rename(columns={
            "sale_number": "# de venta",
            "folio":       "Folio",
            "descripcion": "Descripción",
            "cant":        "Cant.",
            "estado":      "Estado",
        }),
        use_container_width=True,
        height=min(600, 55 + len(df_res) * 35),
    )

    # Download button
    ts     = datetime.now().strftime("%Y%m%d_%H%M%S")
    suffix = "_PRUEBA" if test_mode else ""
    st.download_button(
        label="⬇️ Descargar Excel con resultados",
        data=output_bytes,
        file_name=f"Facturas_Resultado{suffix}_{ts}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
        use_container_width=True,
    )


if __name__ == "__main__":
    main()
